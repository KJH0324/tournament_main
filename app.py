import subprocess
import sys


# 파이썬 버전 확인
def check_python_version(min_version=(3, 11)):
    current_version = sys.version_info
    if current_version < min_version:
        print(f"파이썬 버전이 너무 낮습니다: 현재 버전 {current_version.major}.{current_version.minor} - 최소 {min_version[0]}.{min_version[1]} 필요")
    else:
        print(f"파이썬 버전이 적절합니다: {current_version.major}.{current_version.minor}")

check_python_version()

# 필요한 패키지 목록
required_packages = [
    "python-dotenv", "discord.py", "openpyxl", "pillow"
]

# 패키지 설치 또는 강제 업그레이드 함수
def force_install(package):
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", package])
        print(f"{package} 설치 또는 업데이트 완료.")
    except subprocess.CalledProcessError as e:
        print(f"{package} 설치 실패: {e}")

for package in required_packages:
    print(f"{package} 설치 또는 업데이트 중...")
    force_install(package)


import asyncio
import csv
import io
import json
import logging
import os
import random
import re
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import aiohttp
import discord
from discord import app_commands
from discord.ext import commands
from dotenv import load_dotenv
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

load_dotenv()

log_level = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(level=log_level, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logging.getLogger("discord").setLevel(log_level)
if os.getenv("DEBUG_DISCORD_HTTP") == "1":
    logging.getLogger("discord.http").setLevel(logging.DEBUG)
logger = logging.getLogger(__name__)

TOKEN = os.getenv("DISCORD_BOT_TOKEN")
if not TOKEN:
    raise RuntimeError("DISCORD_BOT_TOKEN is not set. Please configure it in the environment.")

GUILD_ID = 1457245460799553620
BUG_CHANNEL_ID = 1457245461642739838
REPORT_CHANNEL_ID = 1457301492334727304
EMERGENCY_CHANNEL_ID = 1457301552870981722
LOG_CHANNEL_ID = 1457246726766465136
OWNER_ID = 490060722794004480
TOURNAMENT_GUILD_ID = 1194868040895049728
OPEN_TICKET_CATEGORY_ID = 1194868042279161966
CLOSED_TICKET_CATEGORY_ID = 1194868043566813190
TICKET_LOG_CHANNEL_ID = 1458799497357299775
TOURNAMENT_EDIT_ROLE_ID = 1194868041108951115
STAFF_RESIGN_ROLE_ID = 1194868040970547207
COIN_IMAGE_DIR = Path(__file__).parent / "coin"
CHALLONGE_CLIENT_ID = os.getenv("CHALLONGE_CLIENT_ID")
CHALLONGE_CLIENT_SECRET = os.getenv("CHALLONGE_CLIENT_SECRET")
CHALLONGE_API_KEY = os.getenv("CHALLONGE_API_KEY")
CHALLONGE_API_BASE = os.getenv("CHALLONGE_API_BASE", "https://api.challonge.com/v1")
CHALLONGE_TOKEN_URL = os.getenv("CHALLONGE_TOKEN_URL", "https://api.challonge.com/oauth/token")
KST = timezone(timedelta(hours=9))
KST_FONT_URL = "https://github.com/google/fonts/raw/main/ofl/dohyeon/DoHyeon-Regular.ttf"
KST_FONT_PATH = Path(__file__).parent / "data" / "DoHyeon-Regular.ttf"

INTRO_EMBED = discord.Embed(
    title="크즈흐 봇",
    description=(
        "**규정**\n"
        "1. 욕설/비방은 금지됩니다.\n"
        "2. 정확한 상황 설명을 부탁드립니다.\n"
        "3. 비상호출 기능을 남용할 시 제재될 수 있습니다.\n"
        "4. 일반 신고의 경우 답변까지 최대 24시간이 소요됩니다."
    ),
    color=discord.Color.blurple(),
)
INTRO_EMBED.set_footer(text="문의는 모두 기록됩니다.")

CATEGORY_LABELS = {
    "bug": "버그 신고",
    "report": "유저 신고",
    "emergency": "비상 호출",
}

TICKET_NUMBER_RE = re.compile(r"(?:^|-)ticket-(?P<number>\d+)-")

CATEGORY_CHANNELS = {
    "bug": BUG_CHANNEL_ID,
    "report": REPORT_CHANNEL_ID,
    "emergency": EMERGENCY_CHANNEL_ID,
}

CATEGORY_EMOJIS = {
    "bug": "🐞",
    "report": "🕵️",
    "emergency": "🚨",
}


@dataclass
class ThreadBinding:
    thread_id: int
    category: str


class ModerationBot(commands.Bot):
    def __init__(self) -> None:
        intents = discord.Intents.default()
        intents.messages = True
        intents.guilds = True
        intents.dm_messages = True
        super().__init__(command_prefix="!", intents=intents)
        self.user_threads: dict[int, ThreadBinding] = {}

    async def setup_hook(self) -> None:
        logger.info("Starting command registry reset and sync.")
        try:
            await clear_all_command_registries()
        except Exception:
            logger.exception("Failed to clear command registries.")
        for guild_id in (GUILD_ID, TOURNAMENT_GUILD_ID):
            try:
                await sync_guild_commands(guild_id)
            except Exception:
                logger.exception("Failed to sync commands for guild %s", guild_id)


bot = ModerationBot()

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
CONFIG_PATH = DATA_DIR / "config.json"
EVENTS_PATH = DATA_DIR / "events.json"
BACKGROUND_DIR = Path(__file__).parent / "background"
COMMAND_LOG_PATH = DATA_DIR / "command_log.txt"
SCHEDULE_LOG_PATH = DATA_DIR / "schedule_log.txt"
CAPTAINS_CSV_PATH = DATA_DIR / "captains.csv"


@dataclass
class BotConfig:
    bot_op_role: Optional[int] = None
    judge_role: Optional[int] = None
    recorder_role: Optional[int] = None
    schedule_channel: Optional[int] = None
    results_channel: Optional[int] = None
    notification_channel: Optional[int] = None
    transcript_channel: Optional[int] = None
    thumbnail_channel: Optional[int] = None
    tour_logo: Optional[str] = None
    challonge_tournament: Optional[str] = None


@dataclass
class EventData:
    title: str
    schedule_message_id: Optional[int] = None
    schedule_channel_id: Optional[int] = None
    scheduled_event_id: Optional[int] = None
    judge_id: Optional[int] = None
    recorder_id: Optional[int] = None
    details: dict[str, Optional[str]] = field(default_factory=dict)


def load_config() -> BotConfig:
    if CONFIG_PATH.exists():
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        return BotConfig(**data)
    return BotConfig()


def save_config(config: BotConfig) -> None:
    CONFIG_PATH.write_text(json.dumps(config.__dict__, ensure_ascii=False, indent=2), encoding="utf-8")


def load_events() -> dict[str, EventData]:
    if EVENTS_PATH.exists():
        raw = json.loads(EVENTS_PATH.read_text(encoding="utf-8"))
        return {key: EventData(**value) for key, value in raw.items()}
    return {}


def save_events(events: dict[str, EventData]) -> None:
    payload = {key: event.__dict__ for key, event in events.items()}
    EVENTS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


bot_config = load_config()
events_store = load_events()


async def sync_guild_commands(guild_id: int) -> list[app_commands.AppCommand]:
    guild = discord.Object(id=guild_id)
    bot.tree.copy_global_to(guild=guild)
    synced = await bot.tree.sync(guild=guild)
    logger.debug("Synced %s commands to guild %s", len(synced), guild_id)
    if not synced:
        logger.warning("No commands synced to guild %s. Check command registration.", guild_id)
    return synced


async def clear_global_command_registry() -> None:
    global_commands = bot.tree.get_commands()
    if not global_commands:
        return
    bot.tree.clear_commands(guild=None)
    await bot.tree.sync()
    for command in global_commands:
        bot.tree.add_command(command)


async def clear_all_command_registries() -> None:
    global_commands = bot.tree.get_commands()
    try:
        bot.tree.clear_commands(guild=None)
        await bot.tree.sync()
    except Exception:
        logger.exception("Failed to clear global commands.")
    for guild_id in (GUILD_ID, TOURNAMENT_GUILD_ID):
        try:
            guild_obj = discord.Object(id=guild_id)
            bot.tree.clear_commands(guild=guild_obj)
            await bot.tree.sync(guild=guild_obj)
        except Exception:
            logger.exception("Failed to clear commands for guild %s", guild_id)
    for command in global_commands:
        bot.tree.add_command(command)


class CategoryView(discord.ui.View):
    def __init__(self, user_id: int):
        super().__init__(timeout=300)
        self.user_id = user_id

    async def interaction_check(self, interaction: discord.Interaction) -> bool:
        if interaction.user.id != self.user_id:
            await interaction.response.send_message("이 버튼은 요청자만 사용할 수 있어요.", ephemeral=True)
            return False
        return True

    @discord.ui.button(label="버그 신고", style=discord.ButtonStyle.primary, emoji=CATEGORY_EMOJIS["bug"], custom_id="category_bug")
    async def bug_button(self, interaction: discord.Interaction, button: discord.ui.Button) -> None:
        await handle_category_selection(interaction, "bug")

    @discord.ui.button(label="유저 신고", style=discord.ButtonStyle.secondary, emoji=CATEGORY_EMOJIS["report"], custom_id="category_report")
    async def report_button(self, interaction: discord.Interaction, button: discord.ui.Button) -> None:
        await handle_category_selection(interaction, "report")

    @discord.ui.button(label="비상 호출", style=discord.ButtonStyle.danger, emoji=CATEGORY_EMOJIS["emergency"], custom_id="category_emergency")
    async def emergency_button(self, interaction: discord.Interaction, button: discord.ui.Button) -> None:
        await handle_category_selection(interaction, "emergency")


async def get_log_channel() -> Optional[discord.TextChannel]:
    guild = bot.get_guild(GUILD_ID)
    if not guild:
        return None
    channel = guild.get_channel(LOG_CHANNEL_ID)
    if isinstance(channel, discord.TextChannel):
        return channel
    return None


def has_op_role(member: discord.Member) -> bool:
    if bot_config.bot_op_role is None:
        return True
    return any(role.id == bot_config.bot_op_role for role in member.roles)


def has_tournament_edit_role(member: discord.Member) -> bool:
    return any(role.id == TOURNAMENT_EDIT_ROLE_ID for role in member.roles)


def can_edit_event(member: discord.Member, event: EventData) -> bool:
    if has_tournament_edit_role(member):
        return True
    if not has_op_role(member):
        return False
    if event.details.get("result_recorded_at"):
        return False
    start_time = parse_utc_iso(event.details)
    if start_time:
        now = datetime.now(timezone.utc)
        if now >= start_time:
            return False
        if start_time - now <= timedelta(minutes=10):
            return False
    return True


def format_config_value(value: Optional[int], mention_type: str) -> str:
    if value is None:
        return "미설정"
    return f"<{mention_type}{value}>"


def get_tournament_guild() -> Optional[discord.Guild]:
    return bot.get_guild(TOURNAMENT_GUILD_ID)


async def add_member_to_event_channel(member: discord.Member, event: EventData) -> None:
    channel_id = extract_channel_id(event.details.get("channel"))
    if not channel_id:
        return
    channel = member.guild.get_channel(channel_id)
    if not isinstance(channel, discord.TextChannel):
        return
    permissions = channel.overwrites_for(member)
    permissions.view_channel = True
    permissions.send_messages = True
    permissions.read_message_history = True
    await channel.set_permissions(member, overwrite=permissions)


def parse_utc_iso(details: dict[str, Optional[str]]) -> Optional[datetime]:
    utc_iso = details.get("utc_iso")
    if not utc_iso:
        return None
    try:
        return datetime.fromisoformat(utc_iso)
    except ValueError:
        logger.exception("Invalid utc_iso value: %s", utc_iso)
        return None


def parse_time_hm(value: str) -> Optional[tuple[int, int]]:
    try:
        parts = value.strip().split(":")
        if len(parts) != 2:
            return None
        hour = int(parts[0])
        minute = int(parts[1])
        if not (0 <= hour <= 23 and 0 <= minute <= 59):
            return None
        return hour, minute
    except ValueError:
        return None


def append_log_line(path: Path, line: str) -> None:
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    path.parent.mkdir(exist_ok=True)
    with path.open("a", encoding="utf-8", errors="ignore") as handle:
        handle.write(f"{timestamp} | {line}\n")


def log_command_usage(interaction: discord.Interaction, command_name: str) -> None:
    guild_id = interaction.guild_id or "DM"
    channel_id = interaction.channel_id or "unknown"
    user = interaction.user
    append_log_line(
        COMMAND_LOG_PATH,
        f"cmd={command_name} user={user} ({user.id}) guild={guild_id} channel={channel_id}",
    )


def log_schedule_action(
    action: str,
    *,
    user: discord.abc.User,
    event: EventData,
    changes: Optional[list[str]] = None,
) -> None:
    details = event.details
    base = (
        f"{action} title={event.title} match_id={details.get('challonge_match_id', '')} "
        f"user={user} ({user.id})"
    )
    if changes:
        append_log_line(SCHEDULE_LOG_PATH, f"{base} changes={'; '.join(changes)}")
    else:
        summary = (
            f"{base} team1={details.get('team1', '')} team2={details.get('team2', '')} "
            f"utc={details.get('utc_time', '')} local={details.get('local_time', '')}"
        )
        append_log_line(SCHEDULE_LOG_PATH, summary)


def extract_channel_id(raw: Optional[str]) -> Optional[int]:
    if not raw:
        return None
    text = raw.strip()
    if text.startswith("<#") and text.endswith(">"):
        text = text[2:-1]
    try:
        return int(text)
    except ValueError:
        return None


async def ensure_scheduled_event(
    *,
    guild: discord.Guild,
    event: EventData,
    title: str,
    details: dict[str, Optional[str]],
    channel: Optional[discord.TextChannel],
) -> bool:
    if not guild.me or not guild.me.guild_permissions.manage_events:
        logger.warning("Missing Manage Events permission for scheduled event creation.")
        return False
    start_time = parse_utc_iso(details)
    if not start_time:
        logger.warning("Skipping scheduled event sync because utc_iso is missing.")
        return False
    end_time = start_time + timedelta(hours=1)
    if channel is None:
        channel_id = extract_channel_id(details.get("channel"))
        if channel_id:
            channel = guild.get_channel(channel_id)
    location = channel.name if channel else "TBD"
    description = (
        f"Tournament: {details.get('tour_name', '')}\n"
        f"Group: {details.get('group_name', '')}\n"
        f"Round: {details.get('round_no', '')}\n"
        f"Channel: {details.get('channel', '')}"
    ).strip()

    if event.scheduled_event_id:
        try:
            scheduled_event = await guild.fetch_scheduled_event(event.scheduled_event_id)
            await scheduled_event.edit(
                name=title,
                start_time=start_time,
                end_time=end_time,
                description=description or None,
                location=location,
                privacy_level=discord.PrivacyLevel.guild_only,
            )
            return True
        except discord.NotFound:
            logger.warning("Scheduled event %s not found. Recreating.", event.scheduled_event_id)
        except Exception:
            logger.exception("Failed to update scheduled event")
            return False

    try:
        created = await guild.create_scheduled_event(
            name=title,
            start_time=start_time,
            end_time=end_time,
            description=description or None,
            entity_type=discord.EntityType.external,
            location=location,
            privacy_level=discord.PrivacyLevel.guild_only,
        )
        event.scheduled_event_id = created.id
        return True
    except Exception:
        logger.exception("Failed to create scheduled event")
        return False


def build_schedule_embed(title: str, details: dict[str, Optional[str]], event: EventData) -> discord.Embed:
    utc_time = details.get("utc_time", "")
    local_time = details.get("local_time", "")
    tour_name = details.get("tour_name", "")
    group_name = details.get("group_name", "")
    round_no = details.get("round_no", "")
    channel = details.get("channel", "")
    captain1 = details.get("captain1", "")
    captain2 = details.get("captain2", "")
    remarks = details.get("remarks", "")

    embed = discord.Embed(
        title=f"🏆 {title}",
        description=(
            f"UTC Time: {utc_time}\n"
            f"Local Time: {local_time}\n\n"
            f"Tournament: {tour_name}\n"
            f"Group: {group_name}\n"
            f"Round: {round_no}\n\n"
            f"Channel: {channel}\n\n"
            f"Team 1 Captain: {captain1}\n"
            f"Team 2 Captain: {captain2}\n\n"
            f"Remarks: {remarks}"
        ),
        color=discord.Color.red(),
    )

    judge_display = f"<@{event.judge_id}>" if event.judge_id else "미지정"
    recorder_display = f"<@{event.recorder_id}>" if event.recorder_id else "미지정"
    embed.add_field(
        name="Staffs",
        value=f"⚖️ Judge: {judge_display}\n🎥 Recorder: {recorder_display}",
        inline=False,
    )

    return embed


def get_background_image() -> Image.Image:
    if BACKGROUND_DIR.exists():
        images = sorted([path for path in BACKGROUND_DIR.iterdir() if path.suffix.lower() in {".png", ".jpg", ".jpeg"}])
        if images:
            image = Image.open(images[0]).convert("RGB")
            return image.resize((1920, 1080))
    return Image.new("RGB", (1920, 1080), color=(20, 20, 20))


def load_kst_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    try:
        if not KST_FONT_PATH.exists():
            KST_FONT_PATH.parent.mkdir(exist_ok=True)
            urllib.request.urlretrieve(KST_FONT_URL, KST_FONT_PATH)
        return ImageFont.truetype(str(KST_FONT_PATH), size=size)
    except Exception:
        logger.exception("Failed to load KST font; using default.")
        return ImageFont.load_default()


def format_kst_thumbnail_time(details: dict[str, Optional[str]]) -> str:
    utc_iso = details.get("utc_iso")
    if not utc_iso:
        return details.get("utc_time", "") or ""
    try:
        dt_utc = datetime.fromisoformat(utc_iso)
    except ValueError:
        return details.get("utc_time", "") or ""
    dt_kst = dt_utc.astimezone(KST)
    return dt_kst.strftime("%Y-%m-%d %H:%M KST")


def generate_thumbnail(details: dict[str, Optional[str]]) -> discord.File:
    background = get_background_image()
    draw = ImageDraw.Draw(background)
    font_title = load_kst_font(200)
    font_subtitle = load_kst_font(30)

    team1 = details.get("team1", "")
    team2 = details.get("team2", "")

    title_text = f"{team1} vs {team2}".strip()
    time_text = format_kst_thumbnail_time(details)

    title_box = draw.textbbox((0, 0), title_text, font=font_title)
    title_width = title_box[2] - title_box[0]
    title_height = title_box[3] - title_box[1]
    title_x = (background.width - title_width) // 2
    title_y = (background.height - title_height) // 2 - 60
    draw.text((title_x, title_y), title_text, fill=(255, 255, 255), font=font_title)

    time_box = draw.textbbox((0, 0), time_text, font=font_subtitle)
    time_width = time_box[2] - time_box[0]
    time_x = (background.width - time_width) // 2
    time_y = background.height - 140
    draw.text((time_x, time_y), time_text, fill=(220, 220, 220), font=font_subtitle)

    if bot_config.tour_logo:
        try:
            with urllib.request.urlopen(bot_config.tour_logo) as response:
                logo = Image.open(io.BytesIO(response.read())).convert("RGBA")
                logo.thumbnail((300, 300))
                logo_x = (background.width - logo.width) // 2
                background.paste(logo, (logo_x, 40), logo)
        except Exception:
            logger.exception("Failed to load tour logo for thumbnail")

    buffer = io.BytesIO()
    background.save(buffer, format="PNG")
    buffer.seek(0)
    return discord.File(fp=buffer, filename="schedule_thumbnail.png")


def build_results_embed(title: str, details: dict[str, Optional[str]], event: EventData, result_data: dict[str, str]) -> discord.Embed:
    utc_time = details.get("utc_time", "")
    local_time = details.get("local_time", "")
    tour_name = details.get("tour_name", "")
    group_name = details.get("group_name", "")
    round_no = details.get("round_no", "")
    channel = details.get("channel", "")
    captain1 = details.get("captain1", "")
    captain2 = details.get("captain2", "")

    team1 = details.get("team1", "")
    team2 = details.get("team2", "")
    team1_score = result_data.get("team1_score", "")
    team2_score = result_data.get("team2_score", "")
    number_of_matches = result_data.get("number_of_matches", "")
    remarks = result_data.get("remarks", "")
    rec_link = result_data.get("rec_link", "")

    embed = discord.Embed(
        title=f"🗓️ {title}",
        description=(
            f"UTC Time: {utc_time}\n\n"
            f"Local Time: {local_time}\n\n"
            f"Tournament: {tour_name}\n"
            f"Group: {group_name}\n"
            f"Round: {round_no}\n\n"
            f"Channel: {channel}\n\n"
            f"Team1 Captain: {captain1}\n"
            f"Team2 Captain: {captain2}\n\n"
            f"Staffs:\n"
            f"▪️ Judge: {f'<@{event.judge_id}>' if event.judge_id else '미지정'}\n"
            f"▪️ Recorder: {f'<@{event.recorder_id}>' if event.recorder_id else '미지정'}\n\n"
            f"Results:\n"
            f"🏆 {team1} ({team1_score}) : ({team2_score}) {team2} ☠️\n\n"
            f"Remarks: {remarks}\n\n"
            f"Recorder Link: {rec_link}"
        ),
        color=discord.Color.dark_green(),
    )

    screenshot_urls = [value for key, value in result_data.items() if key.startswith("screenshot") and value]
    if screenshot_urls:
        embed.add_field(name="Screenshots", value="\n".join(screenshot_urls), inline=False)
    return embed


class ScheduleView(discord.ui.View):
    def __init__(self, event_title: str):
        super().__init__(timeout=None)
        self.event_title = event_title

    async def interaction_check(self, interaction: discord.Interaction) -> bool:
        if not isinstance(interaction.user, discord.Member):
            await interaction.response.send_message("길드에서만 사용할 수 있어요.")
            return False
        if not has_op_role(interaction.user) and not any(
            role.id in {bot_config.judge_role, bot_config.recorder_role}
            for role in interaction.user.roles
        ):
            await interaction.response.send_message("권한이 없습니다.")
            return False
        return True

    async def update_message(self, interaction: discord.Interaction, event: EventData) -> None:
        embed = build_schedule_embed(event.title, event.details, event)
        self.judge_button.style = (
            discord.ButtonStyle.success if event.judge_id else discord.ButtonStyle.danger
        )
        self.recorder_button.style = (
            discord.ButtonStyle.success if event.recorder_id else discord.ButtonStyle.danger
        )
        await interaction.message.edit(embed=embed, view=self)

    @discord.ui.button(label="Judge", style=discord.ButtonStyle.danger, emoji="⚖️", custom_id="schedule_judge")
    async def judge_button(self, interaction: discord.Interaction, button: discord.ui.Button) -> None:
        event = events_store.get(self.event_title)
        if not event:
            await interaction.response.send_message("이벤트를 찾을 수 없어요.")
            return
        event.judge_id = interaction.user.id
        save_events(events_store)
        if isinstance(interaction.user, discord.Member):
            await add_member_to_event_channel(interaction.user, event)
        await interaction.response.defer()
        await self.update_message(interaction, event)

    @discord.ui.button(label="Recorder", style=discord.ButtonStyle.danger, emoji="🎥", custom_id="schedule_recorder")
    async def recorder_button(self, interaction: discord.Interaction, button: discord.ui.Button) -> None:
        event = events_store.get(self.event_title)
        if not event:
            await interaction.response.send_message("이벤트를 찾을 수 없어요.")
            return
        event.recorder_id = interaction.user.id
        save_events(events_store)
        if isinstance(interaction.user, discord.Member):
            await add_member_to_event_channel(interaction.user, event)
        await interaction.response.defer()
        await self.update_message(interaction, event)


class TicketPanelView(discord.ui.View):
    def __init__(self) -> None:
        super().__init__(timeout=None)

    @discord.ui.button(label="티켓 열기", style=discord.ButtonStyle.success, emoji="🎫", custom_id="ticket_open")
    async def open_ticket(self, interaction: discord.Interaction, button: discord.ui.Button) -> None:
        if interaction.guild_id != TOURNAMENT_GUILD_ID:
            await interaction.response.send_message("토너먼트 서버에서만 사용할 수 있어요.", ephemeral=True)
            return
        if not interaction.guild or not isinstance(interaction.user, discord.Member):
            await interaction.response.send_message("길드에서만 사용할 수 있어요.", ephemeral=True)
            return
        existing = find_existing_ticket_channel(interaction.guild, interaction.user.id)
        if existing:
            await interaction.response.send_message(
                f"이미 열린 티켓이 있습니다: {existing.mention}",
                ephemeral=True,
            )
            return
        category = interaction.guild.get_channel(OPEN_TICKET_CATEGORY_ID)
        if not isinstance(category, discord.CategoryChannel):
            await interaction.response.send_message("열린 티켓 카테고리를 찾을 수 없어요.", ephemeral=True)
            return
        ticket_no = next_ticket_number(interaction.guild)
        base_name = re.sub(r"[^a-z0-9\-]+", "-", interaction.user.display_name.lower()).strip("-")
        base_name = base_name or "ticket"
        channel_name = f"ticket-{ticket_no}-{base_name}-{interaction.user.id}"[:90]
        overwrites = allow_ticket_admins(interaction.guild, interaction.user)
        channel = await interaction.guild.create_text_channel(
            channel_name,
            category=category,
            topic=f"ticket_owner:{interaction.user.id};ticket_no:{ticket_no}",
            overwrites=overwrites,
            reason="Ticket opened",
        )
        await channel.send(
            f"{interaction.user.mention}티켓이 열렸습니다. 관리자가 빠른 시일 내에 답변드릴 예정입니다.\n"
            "이 티켓을 볼 수 있는 관리자(오거나이저 등)에 관한 신고는 봇에게 DM 부탁드립니다."
        )
        await interaction.response.send_message(f"티켓이 생성되었습니다: {channel.mention}", ephemeral=True)


class TicketDeleteView(discord.ui.View):
    def __init__(self) -> None:
        super().__init__(timeout=None)

    @discord.ui.button(label="티켓 삭제", style=discord.ButtonStyle.danger, emoji="🗑️", custom_id="ticket_delete")
    async def delete_ticket(self, interaction: discord.Interaction, button: discord.ui.Button) -> None:
        if not isinstance(interaction.channel, discord.TextChannel) or not interaction.guild:
            await interaction.response.send_message("티켓 채널에서만 사용할 수 있어요.", ephemeral=True)
            return
        if not is_ticket_channel(interaction.channel):
            await interaction.response.send_message("티켓 채널에서만 사용할 수 있어요.", ephemeral=True)
            return
        log_channel = interaction.guild.get_channel(TICKET_LOG_CHANNEL_ID)
        if not isinstance(log_channel, discord.TextChannel):
            await interaction.response.send_message("티켓 로그 채널을 찾을 수 없어요.", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True)
        log_buffer = await build_channel_log(interaction.channel)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"ticket_{interaction.channel.id}_{timestamp}.txt"
        await log_channel.send(
            f"티켓 로그: {interaction.channel.name} ({interaction.channel.id})",
            file=discord.File(log_buffer, filename=filename),
        )
        await interaction.followup.send("티켓 로그를 전송하고 채널을 삭제합니다.", ephemeral=True)
        await interaction.channel.delete(reason="Ticket deleted")


async def create_thread_for_user(user: discord.User, category: str) -> discord.Thread:
    channel_id = CATEGORY_CHANNELS[category]
    guild = bot.get_guild(GUILD_ID)
    if not guild:
        raise RuntimeError("Guild not found")
    channel = guild.get_channel(channel_id)
    if not isinstance(channel, discord.TextChannel):
        raise RuntimeError("Target channel not found")

    thread = await channel.create_thread(
        name=str(user.id),
        type=discord.ChannelType.private_thread,
        reason=f"DM 분류: {CATEGORY_LABELS[category]}",
    )

    mention = f"<@{OWNER_ID}>" if category == "emergency" else ""
    await thread.send(
        f"{CATEGORY_EMOJIS[category]} **{CATEGORY_LABELS[category]}** 요청 접수\n"
        f"유저: {user.mention} ({user.id})\n"
        f"DM 메시지를 아래로 전달받습니다.\n{mention}"
    )
    return thread


async def handle_category_selection(interaction: discord.Interaction, category: str) -> None:
    user = interaction.user
    await interaction.response.defer(ephemeral=True)

    thread = await create_thread_for_user(user, category)
    bot.user_threads[user.id] = ThreadBinding(thread_id=thread.id, category=category)

    log_channel = await get_log_channel()
    if log_channel:
        await log_channel.send(
            f"새 스레드 생성: {thread.mention} | {CATEGORY_LABELS[category]} | 유저 {user} ({user.id})"
        )

    await interaction.followup.send(
        f"분류가 완료되었어요: **{CATEGORY_LABELS[category]}**\n"
        "지금부터 DM으로 보내는 모든 내용은 KJH에게 실시간 전송됩니다.\n"
        "신고자의 익명은 완전히 보호됩니다.",
        ephemeral=True,
    )


async def forward_dm_to_thread(message: discord.Message, thread: discord.Thread) -> None:
    content = message.content or ""
    header = f"**{message.author} ({message.author.id})**"
    payload = f"{header}\n{content}".strip()

    files = []
    for attachment in message.attachments:
        file = await attachment.to_file(use_cached=True)
        files.append(file)

    await thread.send(payload, files=files)


async def send_category_prompt(user: discord.User) -> None:
    view = CategoryView(user.id)
    await user.send(embed=INTRO_EMBED, view=view)


@bot.event
async def on_ready() -> None:
    logger.info("Logged in as %s", bot.user)
    for event in events_store.values():
        if event.schedule_message_id:
            bot.add_view(ScheduleView(event.title))
    bot.add_view(TicketPanelView())
    bot.add_view(TicketDeleteView())


@bot.event
async def on_app_command_completion(
    interaction: discord.Interaction,
    command: app_commands.Command,
) -> None:
    log_command_usage(interaction, command.qualified_name)


@bot.event
async def on_error(event_method: str, *args: object, **kwargs: object) -> None:
    logger.exception("Unhandled exception in %s", event_method)


@bot.tree.error
async def on_app_command_error(interaction: discord.Interaction, error: app_commands.AppCommandError) -> None:
    logger.exception("App command error: %s", error)
    message = "명령 실행 중 오류가 발생했습니다. 콘솔 로그를 확인해 주세요."
    await send_interaction_message(interaction, message, ephemeral=True)


@bot.event
async def on_message(message: discord.Message) -> None:
    if message.author.bot:
        return

    if isinstance(message.channel, discord.DMChannel):
        binding = bot.user_threads.get(message.author.id)
        if not binding:
            await send_category_prompt(message.author)
            return

        thread = bot.get_channel(binding.thread_id)
        if isinstance(thread, discord.Thread):
            await forward_dm_to_thread(message, thread)
        else:
            bot.user_threads.pop(message.author.id, None)
            await send_category_prompt(message.author)
        return

    await bot.process_commands(message)


config_group = app_commands.Group(name="config", description="토너먼트 설정 관리")
events_group = app_commands.Group(name="events", description="토너먼트 이벤트 관리")
general_group = app_commands.Group(name="general", description="일반 관리 도구")
ticket_group = app_commands.Group(name="ticket", description="티켓 관리")
channel_group = app_commands.Group(name="channel", description="채널 관리")
challonge_group = app_commands.Group(name="challonge", description="챌론지 연동 관리")
random_group = app_commands.Group(name="random", description="랜덤 유틸리티")


def format_event_title(team1: str, team2: str) -> str:
    return f"{team1} vs {team2}"


def unique_event_title(team1: str, team2: str, match_id: int) -> str:
    base = format_event_title(team1, team2)
    if base not in events_store:
        return base
    return f"{base} ({match_id})"


async def autocomplete_challonge_teams(
    interaction: discord.Interaction,
    current: str,
) -> list[app_commands.Choice[str]]:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        return []
    tournament_url = bot_config.challonge_tournament
    if not tournament_url:
        return []
    tournament_id = parse_challonge_tournament(tournament_url)
    participants = await fetch_challonge_participants(tournament_id)
    names = []
    for participant in participants:
        name = participant.get("name") or participant.get("display_name")
        if name:
            names.append(name)
    lowered = current.lower()
    filtered = [name for name in names if not lowered or lowered in name.lower()]
    return [app_commands.Choice(name=name, value=name) for name in sorted(filtered)[:25]]


def should_create_match_channel(match: dict) -> bool:
    if match.get("state") == "complete":
        return False
    if match.get("winner_id"):
        return False
    scores = match.get("scores_csv")
    if isinstance(scores, str) and scores.strip():
        return False
    return True


def match_display_name(match: dict, name_by_id: dict[int, str]) -> str:
    player1_id = match.get("player1_id")
    player2_id = match.get("player2_id")
    team1 = name_by_id.get(player1_id, "team1")
    team2 = name_by_id.get(player2_id, "team2")
    round_no = match.get("round")
    round_label = f"R{round_no}" if round_no is not None else "match"
    state = match.get("state") or "unknown"
    match_id = match.get("id")
    match_suffix = f" #{match_id}" if match_id else ""
    return f"{round_label} | {team1} vs {team2} ({state}){match_suffix}"


async def autocomplete_challonge_matches(
    interaction: discord.Interaction,
    current: str,
) -> list[app_commands.Choice[str]]:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        return []
    tournament_url = bot_config.challonge_tournament
    if not tournament_url:
        return []
    tournament_id = parse_challonge_tournament(tournament_url)
    matches = await fetch_challonge_matches(tournament_id)
    participants = await fetch_challonge_participants(tournament_id)
    name_by_id = {
        participant.get("id"): participant.get("name") or participant.get("display_name")
        for participant in participants
    }
    lowered = current.lower()
    choices = []
    for match in matches:
        match_id = match.get("id")
        player1_id = match.get("player1_id")
        player2_id = match.get("player2_id")
        if not match_id or not player1_id or not player2_id:
            continue
        display = match_display_name(match, name_by_id)
        if lowered and lowered not in display.lower():
            continue
        choices.append(app_commands.Choice(name=display, value=str(match_id)))
    return choices[:25]


async def autocomplete_open_challonge_matches(
    interaction: discord.Interaction,
    current: str,
) -> list[app_commands.Choice[str]]:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        return []
    tournament_url = bot_config.challonge_tournament
    if not tournament_url:
        return []
    tournament_id = parse_challonge_tournament(tournament_url)
    matches = await fetch_challonge_matches(tournament_id)
    participants = await fetch_challonge_participants(tournament_id)
    name_by_id = {
        participant.get("id"): participant.get("name") or participant.get("display_name")
        for participant in participants
    }
    lowered = current.lower()
    choices = []
    for match in matches:
        match_id = match.get("id")
        player1_id = match.get("player1_id")
        player2_id = match.get("player2_id")
        if not match_id or not player1_id or not player2_id:
            continue
        if not should_create_match_channel(match):
            continue
        if find_event_by_match_id(str(match_id)):
            continue
        display = match_display_name(match, name_by_id)
        if lowered and lowered not in display.lower():
            continue
        choices.append(app_commands.Choice(name=display, value=str(match_id)))
    return choices[:25]


async def autocomplete_event_matches(
    interaction: discord.Interaction,
    current: str,
    *,
    require_schedule: bool = False,
) -> list[app_commands.Choice[str]]:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        return []
    lowered = current.lower()
    choices = []
    for event in events_store.values():
        match_id = event.details.get("challonge_match_id") if event.details else None
        if not match_id:
            continue
        if require_schedule and not event.schedule_message_id:
            continue
        label = f"{event.title} #{match_id}"
        if lowered and lowered not in label.lower():
            continue
        choices.append(app_commands.Choice(name=label, value=str(match_id)))
    return choices[:25]


async def autocomplete_registered_event_matches(
    interaction: discord.Interaction,
    current: str,
) -> list[app_commands.Choice[str]]:
    return await autocomplete_event_matches(interaction, current, require_schedule=False)


async def autocomplete_scheduled_event_matches(
    interaction: discord.Interaction,
    current: str,
) -> list[app_commands.Choice[str]]:
    return await autocomplete_event_matches(interaction, current, require_schedule=True)


async def autocomplete_staff_resign_roles(
    interaction: discord.Interaction,
    current: str,
) -> list[app_commands.Choice[str]]:
    options = ["Judge", "Recorder"]
    lowered = current.lower()
    return [
        app_commands.Choice(name=option, value=option.lower())
        for option in options
        if not lowered or lowered in option.lower()
    ]


def is_ticket_channel(channel: discord.abc.GuildChannel) -> bool:
    return isinstance(channel, discord.TextChannel) and channel.category_id in {
        OPEN_TICKET_CATEGORY_ID,
        CLOSED_TICKET_CATEGORY_ID,
    }


def extract_ticket_owner_id(channel: discord.TextChannel) -> Optional[int]:
    if not channel.topic:
        return None
    for chunk in channel.topic.split(";"):
        part = chunk.strip()
        if part.startswith("ticket_owner:"):
            value = part.split("ticket_owner:", 1)[1].strip()
            return int(value) if value.isdigit() else None
    return None


def extract_ticket_number(channel: discord.TextChannel) -> Optional[int]:
    if channel.topic:
        for chunk in channel.topic.split(";"):
            part = chunk.strip()
            if part.startswith("ticket_no:"):
                value = part.split("ticket_no:", 1)[1].strip()
                return int(value) if value.isdigit() else None
    match = TICKET_NUMBER_RE.search(channel.name)
    if match:
        return int(match.group("number"))
    return None


def find_existing_ticket_channel(
    guild: discord.Guild,
    user_id: int,
) -> Optional[discord.TextChannel]:
    for channel in guild.channels:
        if not isinstance(channel, discord.TextChannel):
            continue
        if not is_ticket_channel(channel):
            continue
        if extract_ticket_owner_id(channel) == user_id:
            return channel
    return None


def next_ticket_number(guild: discord.Guild) -> int:
    current = 0
    for channel in guild.channels:
        if not isinstance(channel, discord.TextChannel):
            continue
        if not is_ticket_channel(channel):
            continue
        number = extract_ticket_number(channel)
        if number and number > current:
            current = number
    return current + 1


def parse_challonge_tournament(value: str) -> str:
    trimmed = value.strip()
    if "challonge.com" not in trimmed:
        return trimmed
    match = re.search(r"challonge\.com/(?:tournaments/)?(?P<slug>[\w-]+)", trimmed)
    if match:
        return match.group("slug")
    return trimmed.rsplit("/", 1)[-1]


def sanitize_channel_name(value: str) -> str:
    base = re.sub(r"[^a-z0-9가-힣-]+", "-", value.lower()).strip("-")
    return base or "match"


_challonge_token: Optional[str] = None
_challonge_token_expiry: Optional[datetime] = None
_challonge_cache: dict[str, dict[str, object]] = {}


async def get_challonge_token() -> Optional[str]:
    global _challonge_token, _challonge_token_expiry
    if not CHALLONGE_CLIENT_SECRET or not CHALLONGE_CLIENT_ID:
        return None
    if _challonge_token and _challonge_token_expiry:
        if datetime.now(timezone.utc) < _challonge_token_expiry:
            return _challonge_token
    payload = {
        "grant_type": "client_credentials",
        "client_id": CHALLONGE_CLIENT_ID,
        "client_secret": CHALLONGE_CLIENT_SECRET,
    }
    timeout = aiohttp.ClientTimeout(total=15)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        async with session.post(CHALLONGE_TOKEN_URL, data=payload) as response:
            if response.status >= 400:
                body = await response.text()
                logger.error("Failed to fetch Challonge token: %s %s", response.status, body)
                return None
            data = await response.json()
    token = data.get("access_token")
    expires_in = int(data.get("expires_in", 3600))
    if token:
        _challonge_token = token
        _challonge_token_expiry = datetime.now(timezone.utc) + timedelta(seconds=expires_in - 30)
    return token


async def challonge_request(method: str, path: str, *, params: Optional[dict[str, str]] = None, json_body: Optional[dict] = None) -> Optional[dict]:
    headers = {}
    params = params.copy() if params else {}
    if CHALLONGE_API_KEY:
        params["api_key"] = CHALLONGE_API_KEY
        params.setdefault("format", "json")
    else:
        token = await get_challonge_token()
        if not token:
            logger.warning("Challonge token unavailable.")
            return None
        headers["Authorization"] = f"Bearer {token}"
        headers["Accept"] = "application/json"
    url = f"{CHALLONGE_API_BASE}{path}"
    timeout = aiohttp.ClientTimeout(total=20)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        async with session.request(method, url, headers=headers, params=params, json=json_body) as response:
            if response.status >= 400:
                body = await response.text()
                logger.error("Challonge request failed %s %s: %s", method, url, body)
                return None
            try:
                return await response.json(content_type=None)
            except aiohttp.ContentTypeError:
                body = await response.text()
                logger.error("Challonge returned non-JSON payload: %s", body)
                return None


async def fetch_challonge_participants(tournament_id: str) -> list[dict]:
    cache_key = f"{tournament_id}:participants"
    cached = _challonge_cache.get(cache_key)
    now = datetime.now(timezone.utc)
    if cached and isinstance(cached.get("expires_at"), datetime) and cached["expires_at"] > now:
        return cached.get("data", [])
    data = await challonge_request("GET", f"/tournaments/{tournament_id}/participants")
    participants: list[dict]
    if isinstance(data, list):
        participants = data
    elif isinstance(data, dict):
        participants = data.get("participants", [])
    else:
        participants = []
    normalized = []
    for entry in participants:
        if isinstance(entry, dict) and "participant" in entry:
            normalized.append(entry["participant"])
        else:
            normalized.append(entry)
    participants = normalized
    _challonge_cache[cache_key] = {
        "data": participants,
        "expires_at": now + timedelta(minutes=5),
    }
    return participants


async def fetch_challonge_matches(tournament_id: str) -> list[dict]:
    data = await challonge_request("GET", f"/tournaments/{tournament_id}/matches")
    matches: list[dict]
    if isinstance(data, list):
        matches = data
    elif isinstance(data, dict):
        matches = data.get("matches", [])
    else:
        matches = []
    normalized = []
    for entry in matches:
        if isinstance(entry, dict) and "match" in entry:
            normalized.append(entry["match"])
        else:
            normalized.append(entry)
    return normalized


async def fetch_challonge_match(
    tournament_id: str,
    match_id: int,
) -> Optional[dict]:
    matches = await fetch_challonge_matches(tournament_id)
    for match in matches:
        if match.get("id") == match_id:
            return match
    return None


async def report_challonge_result(
    tournament_id: str,
    match_id: int,
    winner_id: int,
    scores_csv: str,
) -> bool:
    payload = {"match": {"winner_id": winner_id, "scores_csv": scores_csv}}
    data = await challonge_request("PUT", f"/tournaments/{tournament_id}/matches/{match_id}", json_body=payload)
    return bool(data)

def clear_challonge_cache() -> None:
    _challonge_cache.clear()
    global _challonge_token, _challonge_token_expiry
    _challonge_token = None
    _challonge_token_expiry = None

def allow_ticket_admins(
    guild: discord.Guild,
    opener: Optional[discord.Member],
) -> dict[discord.Role | discord.Member, discord.PermissionOverwrite]:
    overwrites = {guild.default_role: discord.PermissionOverwrite(view_channel=False)}
    if opener:
        overwrites[opener] = discord.PermissionOverwrite(view_channel=True, send_messages=True, read_message_history=True)
    bot_op_role_id = bot_config.bot_op_role
    if bot_op_role_id:
        role = guild.get_role(bot_op_role_id)
        if role:
            overwrites[role] = discord.PermissionOverwrite(view_channel=True, send_messages=True, read_message_history=True)
    return overwrites


async def build_channel_log(channel: discord.TextChannel) -> io.BytesIO:
    messages = [message async for message in channel.history(limit=None, oldest_first=True)]
    lines = []
    for message in messages:
        timestamp = message.created_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        author = f"{message.author} ({message.author.id})"
        text = message.content or ""
        attachment_lines = [att.url for att in message.attachments]
        combined = "\n".join([text, *attachment_lines]).strip()
        lines.append(f"[{timestamp}] {author}: {combined}")
    content = "\n".join(lines) if lines else "(메시지 없음)"
    buffer = io.BytesIO()
    buffer.write(content.encode("utf-8"))
    buffer.seek(0)
    return buffer


def load_captain_map() -> dict[str, int]:
    if not CAPTAINS_CSV_PATH.exists():
        return {}
    mapping: dict[str, int] = {}
    with CAPTAINS_CSV_PATH.open("r", encoding="utf-8", errors="ignore") as handle:
        sample = handle.read(1024)
        handle.seek(0)
        has_header = "discord" in sample.lower() or "team" in sample.lower() or "name" in sample.lower()
        if has_header:
            reader = csv.DictReader(handle)
            for row in reader:
                team_name = (row.get("team") or row.get("name") or "").strip()
                discord_id = (row.get("discord_id") or row.get("discord") or row.get("id") or "").strip()
                if not team_name or not discord_id:
                    continue
                if discord_id.isdigit():
                    mapping[team_name.lower()] = int(discord_id)
        else:
            reader = csv.reader(handle)
            for row in reader:
                if len(row) < 2:
                    continue
                team_name = row[0].strip()
                discord_id = row[1].strip()
                if not team_name or not discord_id:
                    continue
                if discord_id.isdigit():
                    mapping[team_name.lower()] = int(discord_id)
    return mapping


def save_captain_map(mapping: dict[str, int]) -> None:
    CAPTAINS_CSV_PATH.parent.mkdir(exist_ok=True)
    with CAPTAINS_CSV_PATH.open("w", encoding="utf-8", errors="ignore", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["team", "discord_id"])
        for team_name, discord_id in sorted(mapping.items()):
            writer.writerow([team_name, str(discord_id)])


def load_captain_map_from_xlsx(path: Path) -> dict[str, int]:
    mapping: dict[str, int] = {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return mapping
        header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        has_header = any("team" in cell or "name" in cell or "discord" in cell or "id" == cell for cell in header)
        start_index = 1 if has_header else 0
        team_idx = 0
        id_idx = 1
        if has_header:
            for idx, cell in enumerate(header):
                if "team" in cell or "name" in cell:
                    team_idx = idx
                if "discord" in cell or cell == "id":
                    id_idx = idx
        for row in rows[start_index:]:
            if not row:
                continue
            team_cell = row[team_idx] if team_idx < len(row) else None
            id_cell = row[id_idx] if id_idx < len(row) else None
            if team_cell is None or id_cell is None:
                continue
            team_name = str(team_cell).strip()
            discord_id = str(id_cell).strip()
            if not team_name or not discord_id:
                continue
            if discord_id.isdigit():
                mapping[team_name.lower()] = int(discord_id)
    finally:
        workbook.close()
    return mapping


async def add_captains_to_channel(
    channel: discord.TextChannel,
    team_names: list[str],
    captain_map: dict[str, int],
) -> None:
    for team_name in team_names:
        captain_id = captain_map.get(team_name.lower())
        if not captain_id:
            continue
        member = channel.guild.get_member(captain_id)
        if member is None:
            try:
                member = await channel.guild.fetch_member(captain_id)
            except discord.NotFound:
                continue
            except discord.Forbidden:
                continue
        permissions = channel.overwrites_for(member)
        permissions.view_channel = True
        permissions.send_messages = True
        permissions.read_message_history = True
        await channel.set_permissions(member, overwrite=permissions)


async def build_challonge_match_channels(
    guild: discord.Guild,
    category: discord.CategoryChannel,
    tournament_id: str,
) -> list[discord.TextChannel]:
    matches = await fetch_challonge_matches(tournament_id)
    if not matches:
        return []
    participants = await fetch_challonge_participants(tournament_id)
    name_by_id = {
        participant.get("id"): participant.get("name") or participant.get("display_name")
        for participant in participants
    }
    created = []
    captain_map = load_captain_map()
    existing_names = {channel.name for channel in category.channels if isinstance(channel, discord.TextChannel)}
    existing_match_ids = {
        match_id
        for channel in category.channels
        if isinstance(channel, discord.TextChannel)
        for match_id in [
            int(match.group("id"))
            for match in re.finditer(r"challonge_match_id:(?P<id>\d+)", channel.topic or "")
        ]
    }
    bot_op_role = guild.get_role(bot_config.bot_op_role) if bot_config.bot_op_role else None
    overwrites = {
        guild.default_role: discord.PermissionOverwrite(view_channel=False),
    }
    if bot_op_role:
        overwrites[bot_op_role] = discord.PermissionOverwrite(
            view_channel=True,
            send_messages=True,
            read_message_history=True,
        )
    for match in matches:
        match_id = match.get("id")
        player1_id = match.get("player1_id")
        player2_id = match.get("player2_id")
        if not match_id or not player1_id or not player2_id:
            continue
        if not should_create_match_channel(match):
            continue
        if match_id in existing_match_ids:
            continue
        team1 = name_by_id.get(player1_id, "team1")
        team2 = name_by_id.get(player2_id, "team2")
        round_label = str(match.get("round") or "match")
        base = f"match-{round_label}-{sanitize_channel_name(team1)}-vs-{sanitize_channel_name(team2)}"
        channel_name = base[:90]
        if channel_name in existing_names:
            continue
        channel = await guild.create_text_channel(
            channel_name,
            category=category,
            topic=f"challonge_match_id:{match_id}",
            overwrites=overwrites,
        )
        if captain_map:
            await add_captains_to_channel(channel, [team1, team2], captain_map)
        created.append(channel)
        existing_names.add(channel_name)
    return created


async def send_interaction_message(
    interaction: discord.Interaction,
    message: str,
    *,
    ephemeral: bool = False,
) -> None:
    try:
        if interaction.response.is_done():
            await interaction.followup.send(message, ephemeral=ephemeral)
        else:
            await interaction.response.send_message(message, ephemeral=ephemeral)
    except discord.NotFound:
        logger.warning("Failed to send interaction response because the interaction expired.")


def find_event_by_match_id(match_id: str) -> Optional[tuple[str, EventData]]:
    for title, event in events_store.items():
        stored_match_id = event.details.get("challonge_match_id") if event.details else None
        if stored_match_id and stored_match_id == match_id:
            return title, event
    return None


@config_group.command(name="set", description="토너먼트 봇 사전설정을 저장합니다.")
@app_commands.describe(
    bot_op_role="봇 운영자 역할",
    judge_role="저지 역할",
    recorder_role="레코더 역할",
    schedule_channel="스케줄 채널",
    results_channel="결과 채널",
    notification_channel="알림 채널",
    transcript_channel="트랜스크립트 채널",
    thumbnail_channel="썸네일 채널",
    tour_logo="토너먼트 로고 이미지 URL",
    challonge_tournament="Challonge 토너먼트 링크 또는 ID",
)
async def config_set(
    interaction: discord.Interaction,
    bot_op_role: Optional[discord.Role] = None,
    judge_role: Optional[discord.Role] = None,
    recorder_role: Optional[discord.Role] = None,
    schedule_channel: Optional[discord.TextChannel] = None,
    results_channel: Optional[discord.TextChannel] = None,
    notification_channel: Optional[discord.TextChannel] = None,
    transcript_channel: Optional[discord.TextChannel] = None,
    thumbnail_channel: Optional[discord.TextChannel] = None,
    tour_logo: Optional[str] = None,
    challonge_tournament: Optional[str] = None,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return

    if bot_op_role:
        bot_config.bot_op_role = bot_op_role.id
    if judge_role:
        bot_config.judge_role = judge_role.id
    if recorder_role:
        bot_config.recorder_role = recorder_role.id
    if schedule_channel:
        bot_config.schedule_channel = schedule_channel.id
    if results_channel:
        bot_config.results_channel = results_channel.id
    if notification_channel:
        bot_config.notification_channel = notification_channel.id
    if transcript_channel:
        bot_config.transcript_channel = transcript_channel.id
    if thumbnail_channel:
        bot_config.thumbnail_channel = thumbnail_channel.id
    if tour_logo:
        bot_config.tour_logo = tour_logo
    if challonge_tournament:
        bot_config.challonge_tournament = challonge_tournament

    save_config(bot_config)
    await interaction.response.send_message("설정을 저장했습니다.")


@config_group.command(name="show", description="사전설정을 확인합니다.")
async def config_show(interaction: discord.Interaction) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return

    embed = discord.Embed(title="토너먼트 설정", color=discord.Color.blurple())
    embed.add_field(name="bot_op_role", value=format_config_value(bot_config.bot_op_role, "@&"), inline=False)
    embed.add_field(name="judge_role", value=format_config_value(bot_config.judge_role, "@&"), inline=False)
    embed.add_field(name="recorder_role", value=format_config_value(bot_config.recorder_role, "@&"), inline=False)
    embed.add_field(name="schedule_channel", value=format_config_value(bot_config.schedule_channel, "#"), inline=False)
    embed.add_field(name="results_channel", value=format_config_value(bot_config.results_channel, "#"), inline=False)
    embed.add_field(name="notification_channel", value=format_config_value(bot_config.notification_channel, "#"), inline=False)
    embed.add_field(name="transcript_channel", value=format_config_value(bot_config.transcript_channel, "#"), inline=False)
    embed.add_field(name="thumbnail_channel", value=format_config_value(bot_config.thumbnail_channel, "#"), inline=False)
    embed.add_field(name="tour_logo", value=bot_config.tour_logo or "미설정", inline=False)
    embed.add_field(
        name="challonge_tournament",
        value=bot_config.challonge_tournament or "미설정",
        inline=False,
    )
    await interaction.response.send_message(embed=embed)


@events_group.command(name="create", description="토너먼트 이벤트를 생성합니다.")
@app_commands.describe(
    match="챌론지 매치",
    dd="일",
    mm="월",
    yyyy="연도",
    hour="시",
    minute="분",
    tour_name="토너먼트 이름",
    group_name="그룹",
    round_no="라운드",
    channel="경기 채널",
    captain1="팀 1 캡틴",
    captain2="팀 2 캡틴",
    judge="저지",
    recorder="레코더",
    image_url="썸네일 이미지 URL",
    remarks="비고",
)
@app_commands.autocomplete(match=autocomplete_open_challonge_matches)
async def events_create(
    interaction: discord.Interaction,
    match: str,
    dd: int,
    mm: int,
    yyyy: int,
    hour: int,
    minute: int,
    tour_name: Optional[str] = None,
    group_name: Optional[str] = None,
    round_no: Optional[str] = None,
    channel: Optional[discord.TextChannel] = None,
    captain1: Optional[discord.Member] = None,
    captain2: Optional[discord.Member] = None,
    judge: Optional[discord.Member] = None,
    recorder: Optional[discord.Member] = None,
    image_url: Optional[str] = None,
    remarks: Optional[str] = None,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not bot_config.schedule_channel:
        await interaction.response.send_message("schedule_channel 설정이 필요합니다.")
        return
    if not bot_config.challonge_tournament:
        await interaction.response.send_message("challonge_tournament 설정이 필요합니다.")
        return

    await interaction.response.defer()
    tournament_id = parse_challonge_tournament(bot_config.challonge_tournament)
    try:
        match_id = int(match)
    except ValueError:
        await send_interaction_message(interaction, "유효한 매치를 선택해 주세요.")
        return
    match_data = await fetch_challonge_match(tournament_id, match_id)
    if not match_data:
        await send_interaction_message(interaction, "챌론지 매치를 찾을 수 없어요.")
        return
    if not should_create_match_channel(match_data):
        await send_interaction_message(interaction, "이미 종료된 매치입니다.")
        return
    if find_event_by_match_id(str(match_id)):
        await send_interaction_message(interaction, "이미 등록된 매치입니다.")
        return
    participants = await fetch_challonge_participants(tournament_id)
    name_by_id = {
        participant.get("id"): participant.get("name") or participant.get("display_name")
        for participant in participants
    }
    player1_id = match_data.get("player1_id")
    player2_id = match_data.get("player2_id")
    if not player1_id or not player2_id:
        await send_interaction_message(interaction, "매치 참가 팀이 아직 확정되지 않았어요.")
        return
    team1 = name_by_id.get(player1_id, "team1")
    team2 = name_by_id.get(player2_id, "team2")
    dt_utc = datetime(yyyy, mm, dd, hour, minute, tzinfo=timezone.utc)
    utc_time = dt_utc.strftime("%Y-%m-%d %H:%M")
    local_time = f"{dt_utc.astimezone(KST).strftime('%B %d, %Y %I:%M %p')} ({discord.utils.format_dt(dt_utc, style='R')})"

    title = unique_event_title(team1, team2, match_id)
    details = {
        "team1": team1,
        "team2": team2,
        "utc_time": utc_time,
        "utc_iso": dt_utc.isoformat(),
        "local_time": local_time,
        "tour_name": tour_name or "",
        "group_name": group_name or "",
        "round_no": round_no or "",
        "channel": channel.mention if channel else "unknown",
        "captain1": captain1.mention if captain1 else "unknown",
        "captain2": captain2.mention if captain2 else "unknown",
        "image_url": image_url or "",
        "remarks": remarks or "",
        "challonge_match_id": str(match_id),
        "challonge_player1_id": str(player1_id),
        "challonge_player2_id": str(player2_id),
    }

    event = EventData(
        title=title,
        judge_id=judge.id if judge else None,
        recorder_id=recorder.id if recorder else None,
        details=details,
    )

    tournament_guild = get_tournament_guild()
    schedule_channel = tournament_guild.get_channel(bot_config.schedule_channel) if tournament_guild else None
    if not isinstance(schedule_channel, discord.TextChannel):
        await send_interaction_message(interaction, "스케줄 채널을 찾을 수 없어요.")
        return

    view = ScheduleView(title)
    thumbnail_file = generate_thumbnail(details)
    embed = build_schedule_embed(title, details, event)
    embed.set_thumbnail(url="attachment://schedule_thumbnail.png")
    message = await schedule_channel.send(embed=embed, view=view, file=thumbnail_file)
    event.schedule_message_id = message.id
    event.schedule_channel_id = schedule_channel.id
    scheduled_ok = await ensure_scheduled_event(
        guild=tournament_guild,
        event=event,
        title=title,
        details=details,
        channel=channel,
    )
    events_store[title] = event
    save_events(events_store)
    log_schedule_action("create", user=interaction.user, event=event)

    response = "이벤트를 생성했습니다."
    if not scheduled_ok:
        response += " (디스코드 일정 이벤트 생성에 실패했습니다. 권한을 확인해주세요.)"
    await send_interaction_message(interaction, response)


@events_group.command(name="edit", description="토너먼트 이벤트를 수정합니다.")
@app_commands.describe(
    match="챌론지 매치",
    dd="일",
    mm="월",
    yyyy="연도",
    hour="시",
    minute="분",
    tour_name="토너먼트 이름",
    group_name="그룹",
    round_no="라운드",
    channel="경기 채널",
    captain1="팀 1 캡틴",
    captain2="팀 2 캡틴",
    judge="저지",
    recorder="레코더",
    image_url="썸네일 이미지 URL",
    remarks="비고",
)
@app_commands.autocomplete(match=autocomplete_registered_event_matches)
async def events_edit(
    interaction: discord.Interaction,
    match: str,
    dd: Optional[int] = None,
    mm: Optional[int] = None,
    yyyy: Optional[int] = None,
    hour: Optional[int] = None,
    minute: Optional[int] = None,
    tour_name: Optional[str] = None,
    group_name: Optional[str] = None,
    round_no: Optional[str] = None,
    channel: Optional[discord.TextChannel] = None,
    captain1: Optional[discord.Member] = None,
    captain2: Optional[discord.Member] = None,
    judge: Optional[discord.Member] = None,
    recorder: Optional[discord.Member] = None,
    image_url: Optional[str] = None,
    remarks: Optional[str] = None,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not has_op_role(interaction.user) and not has_tournament_edit_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    event_entry = find_event_by_match_id(match)
    if not event_entry:
        await interaction.response.send_message("이벤트를 찾을 수 없어요.")
        return
    title, event = event_entry
    if not can_edit_event(interaction.user, event):
        await interaction.response.send_message("이 이벤트는 수정할 수 없습니다.")
        return

    before_details = dict(event.details)
    before_judge = event.judge_id
    before_recorder = event.recorder_id
    details = event.details
    if dd and mm and yyyy and hour is not None and minute is not None:
        dt_utc = datetime(yyyy, mm, dd, hour, minute, tzinfo=timezone.utc)
        details["utc_time"] = dt_utc.strftime("%Y-%m-%d %H:%M")
        details["utc_iso"] = dt_utc.isoformat()
        details["local_time"] = (
            f"{dt_utc.astimezone(KST).strftime('%B %d, %Y %I:%M %p')} ({discord.utils.format_dt(dt_utc, style='R')})"
        )
    if tour_name is not None:
        details["tour_name"] = tour_name
    if group_name is not None:
        details["group_name"] = group_name
    if round_no is not None:
        details["round_no"] = round_no
    if channel is not None:
        details["channel"] = channel.mention
    if captain1 is not None:
        details["captain1"] = captain1.mention
    if captain2 is not None:
        details["captain2"] = captain2.mention
    if image_url is not None:
        details["image_url"] = image_url
    if remarks is not None:
        details["remarks"] = remarks
    if judge:
        event.judge_id = judge.id
    if recorder:
        event.recorder_id = recorder.id

    event.details = details
    save_events(events_store)

    await interaction.response.defer()

    if event.schedule_channel_id and event.schedule_message_id:
        tournament_guild = get_tournament_guild()
        channel_obj = tournament_guild.get_channel(event.schedule_channel_id) if tournament_guild else None
        if isinstance(channel_obj, discord.TextChannel):
            try:
                message = await channel_obj.fetch_message(event.schedule_message_id)
                thumbnail_file = generate_thumbnail(details)
                embed = build_schedule_embed(event.title, details, event)
                embed.set_thumbnail(url="attachment://schedule_thumbnail.png")
                await message.edit(embed=embed, view=ScheduleView(event.title), attachments=[thumbnail_file])
            except discord.NotFound:
                pass

    tournament_guild = get_tournament_guild()
    if tournament_guild:
        await ensure_scheduled_event(
            guild=tournament_guild,
            event=event,
            title=event.title,
            details=details,
            channel=channel,
        )

    changes = []
    all_keys = set(before_details.keys()) | set(details.keys())
    for key in sorted(all_keys):
        before_value = before_details.get(key)
        after_value = details.get(key)
        if before_value != after_value:
            changes.append(f"{key}: {before_value} -> {after_value}")
    if before_judge != event.judge_id:
        changes.append(f"judge_id: {before_judge} -> {event.judge_id}")
    if before_recorder != event.recorder_id:
        changes.append(f"recorder_id: {before_recorder} -> {event.recorder_id}")
    if changes:
        log_schedule_action("edit", user=interaction.user, event=event, changes=changes)

    await send_interaction_message(interaction, "이벤트를 수정했습니다.")


@events_group.command(name="delete", description="토너먼트 이벤트를 삭제합니다.")
@app_commands.describe(match="챌론지 매치", reason="삭제 사유")
@app_commands.autocomplete(match=autocomplete_registered_event_matches)
async def events_delete(interaction: discord.Interaction, match: str, reason: Optional[str] = None) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    await interaction.response.defer()
    event_entry = find_event_by_match_id(match)
    if not event_entry:
        await send_interaction_message(interaction, "이벤트를 찾을 수 없어요.")
        return
    event_title, event = event_entry
    events_store.pop(event_title, None)

    save_events(events_store)
    log_schedule_action("delete", user=interaction.user, event=event, changes=[f"reason: {reason or '없음'}"])
    if event.schedule_channel_id and event.schedule_message_id:
        tournament_guild = get_tournament_guild()
        channel_obj = tournament_guild.get_channel(event.schedule_channel_id) if tournament_guild else None
        if isinstance(channel_obj, discord.TextChannel):
            try:
                message = await channel_obj.fetch_message(event.schedule_message_id)
                await message.delete()
            except discord.NotFound:
                pass

    tournament_guild = get_tournament_guild()
    if tournament_guild and event.scheduled_event_id:
        try:
            scheduled_event = await tournament_guild.fetch_scheduled_event(event.scheduled_event_id)
            await scheduled_event.delete()
        except discord.NotFound:
            logger.warning("Scheduled event %s not found for deletion.", event.scheduled_event_id)
        except Exception:
            logger.exception("Failed to delete scheduled event")

    await send_interaction_message(interaction, f"이벤트를 삭제했습니다. 사유: {reason or '없음'}")


@events_group.command(name="list", description="저장된 토너먼트 이벤트 목록을 확인합니다.")
async def events_list(interaction: discord.Interaction) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return

    if not events_store:
        await interaction.response.send_message("저장된 이벤트가 없습니다.")
        return

    titles = "\n".join(f"- {title}" for title in sorted(events_store.keys()))
    embed = discord.Embed(title="저장된 이벤트 목록", description=titles, color=discord.Color.blurple())
    await interaction.response.send_message(embed=embed)


@events_group.command(name="show", description="스케줄 임베드를 다시 표시합니다.")
@app_commands.describe(match="챌론지 매치")
@app_commands.autocomplete(match=autocomplete_scheduled_event_matches)
async def events_show(
    interaction: discord.Interaction,
    match: str,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not (has_op_role(interaction.user) or any(role.id == STAFF_RESIGN_ROLE_ID for role in interaction.user.roles)):
        await interaction.response.send_message("권한이 없습니다.")
        return
    event_entry = find_event_by_match_id(match)
    if not event_entry:
        await interaction.response.send_message("이벤트를 찾을 수 없어요.")
        return
    _, event_data = event_entry
    if not event_data.schedule_message_id:
        await interaction.response.send_message("스케줄이 생성된 이벤트만 표시할 수 있어요.")
        return
    thumbnail_file = generate_thumbnail(event_data.details)
    embed = build_schedule_embed(event_data.title, event_data.details, event_data)
    embed.set_thumbnail(url="attachment://schedule_thumbnail.png")
    await interaction.response.send_message(embed=embed, file=thumbnail_file)


@events_group.command(name="results", description="토너먼트 결과를 등록합니다.")
@app_commands.describe(
    event="챌론지 매치",
    team1_score="팀 1 점수",
    team2_score="팀 2 점수",
    number_of_matches="매치 수",
    remarks="비고",
    rec_link="레코더 링크",
    screenshot1="스크린샷 URL 1",
    screenshot2="스크린샷 URL 2",
    screenshot3="스크린샷 URL 3",
    screenshot4="스크린샷 URL 4",
    screenshot5="스크린샷 URL 5",
    screenshot6="스크린샷 URL 6",
    screenshot7="스크린샷 URL 7",
    screenshot8="스크린샷 URL 8",
    screenshot9="스크린샷 URL 9",
    screenshot10="스크린샷 URL 10",
)
@app_commands.autocomplete(event=autocomplete_scheduled_event_matches)
async def events_results(
    interaction: discord.Interaction,
    event: str,
    team1_score: str,
    team2_score: str,
    number_of_matches: str,
    remarks: Optional[str] = None,
    rec_link: Optional[str] = None,
    screenshot1: Optional[str] = None,
    screenshot2: Optional[str] = None,
    screenshot3: Optional[str] = None,
    screenshot4: Optional[str] = None,
    screenshot5: Optional[str] = None,
    screenshot6: Optional[str] = None,
    screenshot7: Optional[str] = None,
    screenshot8: Optional[str] = None,
    screenshot9: Optional[str] = None,
    screenshot10: Optional[str] = None,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not (has_op_role(interaction.user) or any(role.id == STAFF_RESIGN_ROLE_ID for role in interaction.user.roles)):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not bot_config.results_channel:
        await interaction.response.send_message("results_channel 설정이 필요합니다.")
        return

    event_entry = find_event_by_match_id(event)
    if not event_entry:
        await interaction.response.send_message("이벤트를 찾을 수 없어요.")
        return
    _, event_data = event_entry
    if not event_data.schedule_message_id:
        await interaction.response.send_message("스케줄이 생성된 이벤트만 결과를 등록할 수 있어요.")
        return

    result_payload = {
        "team1_score": team1_score,
        "team2_score": team2_score,
        "number_of_matches": number_of_matches,
        "remarks": remarks or "",
        "rec_link": rec_link or "",
        "screenshot1": screenshot1 or "",
        "screenshot2": screenshot2 or "",
        "screenshot3": screenshot3 or "",
        "screenshot4": screenshot4 or "",
        "screenshot5": screenshot5 or "",
        "screenshot6": screenshot6 or "",
        "screenshot7": screenshot7 or "",
        "screenshot8": screenshot8 or "",
        "screenshot9": screenshot9 or "",
        "screenshot10": screenshot10 or "",
    }

    tournament_guild = get_tournament_guild()
    results_channel = tournament_guild.get_channel(bot_config.results_channel) if tournament_guild else None
    if not isinstance(results_channel, discord.TextChannel):
        await interaction.response.send_message("결과 채널을 찾을 수 없어요.")
        return

    embed = build_results_embed(event_data.title, event_data.details, event_data, result_payload)
    await results_channel.send(embed=embed)
    event_data.details["result_recorded_at"] = datetime.now(timezone.utc).isoformat()
    save_events(events_store)
    match_id_raw = event_data.details.get("challonge_match_id") if event_data.details else None
    if match_id_raw and bot_config.challonge_tournament:
        try:
            match_id = int(match_id_raw)
            player1_id = int(event_data.details.get("challonge_player1_id") or 0)
            player2_id = int(event_data.details.get("challonge_player2_id") or 0)
            team1_score_int = int(team1_score)
            team2_score_int = int(team2_score)
            if team1_score_int == team2_score_int:
                logger.warning("Match result is tied; skipping Challonge update.")
                winner_id = None
            else:
                winner_id = player1_id if team1_score_int > team2_score_int else player2_id
            scores_csv = f"{team1_score_int}-{team2_score_int}"
            tournament_id = parse_challonge_tournament(bot_config.challonge_tournament)
            if winner_id:
                updated = await report_challonge_result(tournament_id, match_id, winner_id, scores_csv)
                if not updated:
                    logger.warning("Failed to update Challonge match %s", match_id)
        except (ValueError, TypeError):
            logger.warning("Invalid Challonge match data; skipping auto update.")
    await interaction.response.send_message("결과를 등록했습니다.")


@events_group.command(name="staff_resign", description="스태프 역할을 포기합니다.")
@app_commands.describe(match="챌론지 매치", reason="사유", role="judge 또는 recorder")
@app_commands.autocomplete(match=autocomplete_registered_event_matches, role=autocomplete_staff_resign_roles)
async def events_staff_resign(
    interaction: discord.Interaction,
    match: str,
    reason: Optional[str] = None,
    role: Optional[str] = None,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not any(role.id == STAFF_RESIGN_ROLE_ID for role in interaction.user.roles):
        await interaction.response.send_message("권한이 없습니다.")
        return
    await interaction.response.defer()
    event_entry = find_event_by_match_id(match)
    if not event_entry:
        await send_interaction_message(interaction, "이벤트를 찾을 수 없어요.")
        return
    _, event = event_entry

    role_key = (role or "").lower()
    if role_key == "judge":
        event.judge_id = None
    elif role_key == "recorder":
        event.recorder_id = None
    else:
        await interaction.response.send_message("role 파라미터는 judge 또는 recorder 이어야 합니다.")
        return

    save_events(events_store)
    if event.schedule_channel_id and event.schedule_message_id:
        tournament_guild = get_tournament_guild()
        channel_obj = tournament_guild.get_channel(event.schedule_channel_id) if tournament_guild else None
        if isinstance(channel_obj, discord.TextChannel):
            try:
                message = await channel_obj.fetch_message(event.schedule_message_id)
                thumbnail_file = generate_thumbnail(event.details)
                embed = build_schedule_embed(event.title, event.details, event)
                embed.set_thumbnail(url="attachment://schedule_thumbnail.png")
                await message.edit(embed=embed, view=ScheduleView(event.title), attachments=[thumbnail_file])
            except discord.NotFound:
                pass

    await send_interaction_message(interaction, f"스태프 역할을 포기했습니다. 사유: {reason or '없음'}")


@events_group.command(name="reset_tournament", description="모든 토너먼트 정보를 초기화합니다.")
async def events_reset_tournament(interaction: discord.Interaction) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_tournament_edit_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    events_store.clear()
    save_events(events_store)
    await interaction.response.send_message("토너먼트 정보를 초기화했습니다.")


@general_group.command(name="add_to_channel", description="멤버 또는 역할에 채널 권한을 부여합니다.")
@app_commands.describe(
    member_role="멤버 또는 역할",
    channel="채널 (비워두면 현재 채널)",
    send_messages="메시지 전송",
    read_message_history="메시지 기록 읽기",
)
async def general_add_to_channel(
    interaction: discord.Interaction,
    member_role: discord.Member | discord.Role,
    channel: Optional[discord.TextChannel] = None,
    send_messages: bool = True,
    read_message_history: bool = True,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    target_channel = channel
    if target_channel is None:
        if isinstance(interaction.channel, discord.TextChannel):
            target_channel = interaction.channel
        else:
            await interaction.response.send_message("채널을 찾을 수 없어요.")
            return
    permissions = target_channel.overwrites_for(member_role)
    permissions.view_channel = True
    permissions.send_messages = send_messages
    permissions.read_message_history = read_message_history
    await target_channel.set_permissions(member_role, overwrite=permissions)
    await interaction.response.send_message("권한을 추가했습니다.")


@general_group.command(name="remove_from_channel", description="멤버 또는 역할의 채널 권한을 제거합니다.")
@app_commands.describe(
    member_role="멤버 또는 역할",
    channel="채널 (비워두면 현재 채널)",
)
async def general_remove_from_channel(
    interaction: discord.Interaction,
    member_role: discord.Member | discord.Role,
    channel: Optional[discord.TextChannel] = None,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    target_channel = channel
    if target_channel is None:
        if isinstance(interaction.channel, discord.TextChannel):
            target_channel = interaction.channel
        else:
            await interaction.response.send_message("채널을 찾을 수 없어요.")
            return
    await target_channel.set_permissions(member_role, overwrite=None)
    await interaction.response.send_message("권한을 제거했습니다.")


@general_group.command(name="close_channel", description="채널을 닫고 로그를 전송합니다.")
@app_commands.describe(channel="닫을 채널", save_transcript="트랜스크립트 저장 여부")
async def general_close_channel(
    interaction: discord.Interaction,
    channel: discord.TextChannel,
    save_transcript: bool = True,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    transcript_channel_id = bot_config.transcript_channel
    if save_transcript and not transcript_channel_id:
        await interaction.response.send_message("transcript_channel 설정이 필요합니다.")
        return

    if save_transcript and transcript_channel_id:
        tournament_guild = get_tournament_guild()
        transcript_channel = tournament_guild.get_channel(transcript_channel_id) if tournament_guild else None
        if isinstance(transcript_channel, discord.TextChannel):
            messages = [message async for message in channel.history(limit=None, oldest_first=True)]
            lines = []
            for message in messages:
                timestamp = message.created_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
                author = f"{message.author} ({message.author.id})"
                text = message.content or ""
                attachment_lines = [att.url for att in message.attachments]
                combined = "\n".join([text, *attachment_lines]).strip()
                lines.append(f"[{timestamp}] {author}: {combined}")
            content = "\n".join(lines) if lines else "(메시지 없음)"
            buffer = io.BytesIO(content.encode("utf-8"))
            buffer.seek(0)
            await transcript_channel.send(
                f"채널 로그: {channel.name} ({channel.id})",
                file=discord.File(buffer, filename=f"channel_{channel.id}.txt"),
            )

    bot_op_role = channel.guild.get_role(bot_config.bot_op_role) if bot_config.bot_op_role else None
    new_overwrites = {channel.guild.default_role: discord.PermissionOverwrite(view_channel=False)}
    if bot_op_role:
        new_overwrites[bot_op_role] = discord.PermissionOverwrite(
            view_channel=True,
            send_messages=False,
            read_message_history=True,
        )
    await channel.edit(overwrites=new_overwrites)
    await interaction.response.send_message("채널을 닫았습니다.")


@ticket_group.command(name="panel", description="티켓 생성 패널을 보냅니다.")
async def ticket_panel(interaction: discord.Interaction) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("이 명령은 봇 소유자만 사용할 수 있어요.")
        return
    embed = discord.Embed(
        title="티켓 생성",
        description="아래 버튼을 눌러 티켓을 열어주세요.",
        color=discord.Color.blurple(),
    )
    await interaction.response.send_message(embed=embed, view=TicketPanelView())


@ticket_group.command(name="close", description="티켓을 닫습니다.")
async def ticket_close(interaction: discord.Interaction) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.channel, discord.TextChannel):
        await interaction.response.send_message("티켓 채널에서만 사용할 수 있어요.")
        return
    if not is_ticket_channel(interaction.channel):
        await interaction.response.send_message("티켓 채널에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    await interaction.response.defer(ephemeral=True)

    closed_category = interaction.guild.get_channel(CLOSED_TICKET_CATEGORY_ID)
    opener_id = extract_ticket_owner_id(interaction.channel)
    opener = interaction.guild.get_member(opener_id) if opener_id else None
    ticket_no = extract_ticket_number(interaction.channel) or next_ticket_number(interaction.guild)
    opener_name = (
        re.sub(r"[^a-z0-9\\-]+", "-", opener.display_name.lower()).strip("-")
        if opener
        else "unknown"
    )
    opener_name = opener_name or "unknown"
    new_name = f"closed-{ticket_no}-{opener_name}-{opener_id or 'unknown'}"[:90]
    overwrites = {interaction.guild.default_role: discord.PermissionOverwrite(view_channel=False)}
    if opener:
        overwrites[opener] = discord.PermissionOverwrite(view_channel=False)
    bot_op_role_id = bot_config.bot_op_role
    if bot_op_role_id:
        role = interaction.guild.get_role(bot_op_role_id)
        if role:
            overwrites[role] = discord.PermissionOverwrite(view_channel=True, send_messages=True, read_message_history=True)

    await interaction.channel.edit(
        category=closed_category if isinstance(closed_category, discord.CategoryChannel) else None,
        name=new_name,
        overwrites=overwrites,
        reason="Ticket closed",
    )

    embed = discord.Embed(
        title="티켓이 닫혔습니다.",
        description="필요하면 아래 버튼으로 티켓을 삭제하고 로그를 전송할 수 있어요.",
        color=discord.Color.dark_grey(),
    )
    await interaction.channel.send(embed=embed, view=TicketDeleteView())
    await interaction.followup.send("티켓을 닫았습니다.", ephemeral=True)


@channel_group.command(name="create", description="챌론지 대진표 기반 매치 채널을 생성합니다.")
@app_commands.describe(
    challonge_link="Challonge 토너먼트 링크 또는 ID",
    category="채널을 만들 카테고리",
)
async def channel_create(
    interaction: discord.Interaction,
    challonge_link: str,
    category: Optional[discord.CategoryChannel] = None,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not CHALLONGE_API_KEY and (not CHALLONGE_CLIENT_SECRET or not CHALLONGE_CLIENT_ID):
        await interaction.response.send_message("CHALLONGE_API_KEY 또는 CHALLONGE_CLIENT_ID/SECRET 환경 변수가 필요합니다.")
        return
    await interaction.response.defer(ephemeral=True)
    tournament_id = parse_challonge_tournament(challonge_link)
    target_category = category or (
        interaction.channel.category if isinstance(interaction.channel, discord.TextChannel) else None
    )
    if not isinstance(target_category, discord.CategoryChannel):
        await interaction.followup.send("카테고리를 찾을 수 없어요.", ephemeral=True)
        return
    created = await build_challonge_match_channels(interaction.guild, target_category, tournament_id)
    if not created:
        await interaction.followup.send("생성할 매치 채널이 없습니다.", ephemeral=True)
        return
    bot_config.challonge_tournament = challonge_link
    save_config(bot_config)
    await interaction.followup.send(
        f"매치 채널 {len(created)}개를 생성했습니다.",
        ephemeral=True,
    )


@challonge_group.command(name="create", description="챌론지 매치 기반 채널을 생성합니다.")
@app_commands.describe(category="채널을 만들 카테고리")
async def challonge_create(
    interaction: discord.Interaction,
    category: Optional[discord.CategoryChannel] = None,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not bot_config.challonge_tournament:
        await interaction.response.send_message("challonge_tournament 설정이 필요합니다.")
        return
    if not CHALLONGE_API_KEY and (not CHALLONGE_CLIENT_SECRET or not CHALLONGE_CLIENT_ID):
        await interaction.response.send_message("CHALLONGE_API_KEY 또는 CHALLONGE_CLIENT_ID/SECRET 환경 변수가 필요합니다.")
        return
    await interaction.response.defer()
    tournament_id = parse_challonge_tournament(bot_config.challonge_tournament)
    target_category = category or (
        interaction.channel.category if isinstance(interaction.channel, discord.TextChannel) else None
    )
    if not isinstance(target_category, discord.CategoryChannel):
        await interaction.followup.send("카테고리를 찾을 수 없어요.")
        return
    created = await build_challonge_match_channels(interaction.guild, target_category, tournament_id)
    if not created:
        await interaction.followup.send("생성할 매치 채널이 없습니다.")
        return
    await interaction.followup.send(
        f"매치 채널 {len(created)}개를 생성했습니다.",
    )

@challonge_group.command(name="set", description="챌론지 토너먼트를 설정합니다.")
@app_commands.describe(
    tournament="Challonge 토너먼트 링크 또는 ID",
    captains_csv="팀장 매핑 CSV (선택)",
)
async def challonge_set(
    interaction: discord.Interaction,
    tournament: str,
    captains_csv: Optional[discord.Attachment] = None,
) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if captains_csv:
        filename = captains_csv.filename.lower()
        if filename.endswith(".csv"):
            await captains_csv.save(CAPTAINS_CSV_PATH)
        elif filename.endswith(".xlsx"):
            temp_path = DATA_DIR / "captains_upload.xlsx"
            try:
                await captains_csv.save(temp_path)
                mapping = load_captain_map_from_xlsx(temp_path)
                save_captain_map(mapping)
            finally:
                try:
                    temp_path.unlink()
                except FileNotFoundError:
                    pass
        else:
            await interaction.response.send_message("CSV 또는 XLSX 파일만 업로드할 수 있어요.")
            return
    bot_config.challonge_tournament = tournament
    save_config(bot_config)
    message = "챌론지 토너먼트를 설정했습니다."
    if captains_csv:
        message += " (팀장 CSV를 저장했습니다.)"
    await interaction.response.send_message(message)


@challonge_group.command(name="info", description="챌론지 연동 상태를 확인합니다.")
async def challonge_info(interaction: discord.Interaction) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    embed = discord.Embed(title="챌론지 연동 상태", color=discord.Color.blurple())
    embed.add_field(name="tournament", value=bot_config.challonge_tournament or "미설정", inline=False)
    embed.add_field(
        name="client_secret",
        value="설정됨" if CHALLONGE_CLIENT_SECRET else "미설정",
        inline=False,
    )
    embed.add_field(
        name="client_id",
        value="설정됨" if CHALLONGE_CLIENT_ID else "미설정",
        inline=False,
    )
    embed.add_field(
        name="api_key",
        value="설정됨" if CHALLONGE_API_KEY else "미설정",
        inline=False,
    )
    await interaction.response.send_message(embed=embed)


@challonge_group.command(name="participants", description="챌론지 참가 팀 목록을 확인합니다.")
async def challonge_participants(interaction: discord.Interaction) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not bot_config.challonge_tournament:
        await interaction.response.send_message("challonge_tournament 설정이 필요합니다.")
        return
    if not CHALLONGE_API_KEY and (not CHALLONGE_CLIENT_SECRET or not CHALLONGE_CLIENT_ID):
        await interaction.response.send_message("CHALLONGE_API_KEY 또는 CHALLONGE_CLIENT_ID/SECRET 환경 변수가 필요합니다.")
        return
    await interaction.response.defer()
    tournament_id = parse_challonge_tournament(bot_config.challonge_tournament)
    participants = await fetch_challonge_participants(tournament_id)
    if not participants:
        await interaction.followup.send("참가 팀을 찾을 수 없어요.")
        return
    names = []
    for participant in participants:
        name = participant.get("name") or participant.get("display_name")
        if name:
            names.append(name)
    names_sorted = sorted(names)
    preview = "\n".join(f"- {name}" for name in names_sorted[:50])
    embed = discord.Embed(
        title="챌론지 참가 팀",
        description=preview or "목록이 비어 있습니다.",
        color=discord.Color.blurple(),
    )
    if len(names_sorted) > 50:
        embed.set_footer(text=f"총 {len(names_sorted)}팀 중 50개 표시")
    await interaction.followup.send(embed=embed)


@challonge_group.command(name="matches", description="챌론지 매치 목록을 확인합니다.")
async def challonge_matches(interaction: discord.Interaction) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    if not bot_config.challonge_tournament:
        await interaction.response.send_message("challonge_tournament 설정이 필요합니다.")
        return
    if not CHALLONGE_API_KEY and (not CHALLONGE_CLIENT_SECRET or not CHALLONGE_CLIENT_ID):
        await interaction.response.send_message("CHALLONGE_API_KEY 또는 CHALLONGE_CLIENT_ID/SECRET 환경 변수가 필요합니다.")
        return
    await interaction.response.defer()
    tournament_id = parse_challonge_tournament(bot_config.challonge_tournament)
    matches = await fetch_challonge_matches(tournament_id)
    if not matches:
        await interaction.followup.send("매치 정보를 찾을 수 없어요.")
        return
    participants = await fetch_challonge_participants(tournament_id)
    name_by_id = {
        participant.get("id"): participant.get("name") or participant.get("display_name")
        for participant in participants
    }
    lines = []
    for match in matches[:50]:
        player1_id = match.get("player1_id")
        player2_id = match.get("player2_id")
        if not player1_id or not player2_id:
            continue
        p1 = name_by_id.get(player1_id, "team1")
        p2 = name_by_id.get(player2_id, "team2")
        state = match.get("state") or "unknown"
        score = match.get("scores_csv") or "-"
        lines.append(f"- {p1} vs {p2} ({state}, {score})")
    embed = discord.Embed(
        title="챌론지 매치",
        description="\n".join(lines) or "표시할 매치가 없습니다.",
        color=discord.Color.blurple(),
    )
    if len(matches) > 50:
        embed.set_footer(text=f"총 {len(matches)}경기 중 50개 표시")
    await interaction.followup.send(embed=embed)


@challonge_group.command(name="refresh", description="챌론지 캐시를 초기화합니다.")
async def challonge_refresh(interaction: discord.Interaction) -> None:
    if interaction.guild_id != TOURNAMENT_GUILD_ID:
        await interaction.response.send_message("이 명령은 토너먼트 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.user, discord.Member) or not has_op_role(interaction.user):
        await interaction.response.send_message("권한이 없습니다.")
        return
    clear_challonge_cache()
    await interaction.response.send_message("챌론지 캐시를 초기화했습니다.")


@random_group.command(name="time", description="UTC 기준 30분 단위 랜덤 시간을 뽑습니다.")
@app_commands.describe(start="시작 시간 (HH:MM, UTC)", end="종료 시간 (HH:MM, UTC)")
async def random_time(
    interaction: discord.Interaction,
    start: str,
    end: str,
) -> None:
    start_parts = parse_time_hm(start)
    end_parts = parse_time_hm(end)
    if not start_parts or not end_parts:
        await interaction.response.send_message("시간 형식은 HH:MM (UTC)입니다.")
        return
    start_dt = datetime(2000, 1, 1, start_parts[0], start_parts[1], tzinfo=timezone.utc)
    end_dt = datetime(2000, 1, 1, end_parts[0], end_parts[1], tzinfo=timezone.utc)
    if end_dt <= start_dt:
        await interaction.response.send_message("종료 시간은 시작 시간보다 늦어야 합니다.")
        return
    slots = []
    current = start_dt
    while current <= end_dt:
        slots.append(current)
        current += timedelta(minutes=30)
    if not slots:
        await interaction.response.send_message("해당 범위에서 시간을 찾을 수 없어요.")
        return
    selected = random.choice(slots)
    await interaction.response.send_message(
        f"랜덤 시간: {selected.strftime('%H:%M')} UTC (범위: {start_dt.strftime('%H:%M')} ~ {end_dt.strftime('%H:%M')})"
    )


bot.tree.add_command(config_group)
bot.tree.add_command(events_group)
bot.tree.add_command(general_group)
bot.tree.add_command(ticket_group)
bot.tree.add_command(channel_group)
bot.tree.add_command(challonge_group)
bot.tree.add_command(random_group)


@bot.tree.command(name="toss", description="코인 토스를 합니다.")
async def toss_command(interaction: discord.Interaction) -> None:
    if not COIN_IMAGE_DIR.exists():
        await interaction.response.send_message("코인 이미지 디렉토리를 찾을 수 없어요.")
        return
    candidates = [path for path in COIN_IMAGE_DIR.iterdir() if path.name in {"A.png", "B.png"}]
    if not candidates:
        await interaction.response.send_message("코인 이미지(A.png, B.png)를 찾을 수 없어요.")
        return
    selected = random.choice(candidates)
    embed = discord.Embed(title="코인 토스", color=discord.Color.blurple())
    embed.set_image(url=f"attachment://{selected.name}")
    await interaction.response.send_message(embed=embed, file=discord.File(selected, filename=selected.name))


@bot.tree.command(name="답장", description="스레드에서 DM으로 답장합니다.")
@app_commands.describe(content="전송할 메시지")
async def reply_command(interaction: discord.Interaction, content: str) -> None:
    if interaction.guild_id != GUILD_ID:
        await interaction.response.send_message("이 명령은 봇 운영 서버에서만 사용할 수 있어요.")
        return
    if not isinstance(interaction.channel, discord.Thread):
        await interaction.response.send_message("이 명령은 스레드에서만 사용할 수 있어요.")
        return

    thread = interaction.channel
    try:
        user_id = int(thread.name)
    except ValueError:
        await interaction.response.send_message("스레드 이름에서 유저 ID를 찾을 수 없어요.")
        return

    user = bot.get_user(user_id) or await bot.fetch_user(user_id)
    await user.send(content)

    await thread.send(
        f"📨 **답장 전송**\n"
        f"담당자: {interaction.user.mention}\n"
        f"내용: {content}"
    )
    await interaction.response.send_message("답장을 전송했어요.")


@bot.tree.command(name="sync", description="슬래시 명령어를 즉시 업데이트합니다.")
async def sync_commands(interaction: discord.Interaction) -> None:
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("이 명령은 봇 소유자만 사용할 수 있어요.")
        return

    await interaction.response.defer(ephemeral=True)
    try:
        await clear_all_command_registries()
    except Exception:
        logger.exception("Failed to clear command registries via /sync.")
    updated = []
    for guild_id in (GUILD_ID, TOURNAMENT_GUILD_ID):
        try:
            await sync_guild_commands(guild_id)
            updated.append(str(guild_id))
        except Exception:
            logger.exception("Failed to sync commands for guild %s via /sync", guild_id)

    await interaction.followup.send(
        f"명령어를 업데이트했습니다. (guilds: {', '.join(updated)})",
        ephemeral=True,
    )


async def build_thread_log(thread: discord.Thread) -> io.BytesIO:
    messages = [message async for message in thread.history(limit=None, oldest_first=True)]
    lines = []
    for message in messages:
        timestamp = message.created_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        author = f"{message.author} ({message.author.id})"
        text = message.content or ""
        attachment_lines = [att.url for att in message.attachments]
        combined = "\n".join([text, *attachment_lines]).strip()
        lines.append(f"[{timestamp}] {author}: {combined}")

    content = "\n".join(lines) if lines else "(메시지 없음)"
    buffer = io.BytesIO()
    buffer.write(content.encode("utf-8"))
    buffer.seek(0)
    return buffer


@bot.tree.command(name="닫기", description="스레드를 닫고 로그를 전송합니다.")
async def close_thread(interaction: discord.Interaction) -> None:
    if interaction.guild_id != GUILD_ID:
        await interaction.response.send_message("이 명령은 봇 운영 서버에서만 사용할 수 있어요.", ephemeral=True)
        return
    if not isinstance(interaction.channel, discord.Thread):
        await interaction.response.send_message("이 명령은 스레드에서만 사용할 수 있어요.", ephemeral=True)
        return

    thread = interaction.channel
    await interaction.response.defer(ephemeral=True)

    log_channel = await get_log_channel()
    if not log_channel:
        await interaction.followup.send("로그 채널을 찾을 수 없어요.", ephemeral=True)
        return

    log_buffer = await build_thread_log(thread)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"thread_{thread.id}_{timestamp}.txt"
    await log_channel.send(
        f"스레드 종료 로그: {thread.name} ({thread.id})",
        file=discord.File(log_buffer, filename=filename),
    )

    await thread.edit(archived=True, locked=True)
    await interaction.followup.send("스레드를 닫고 로그를 전송했습니다.", ephemeral=True)

    try:
        user_id = int(thread.name)
    except ValueError:
        return
    bot.user_threads.pop(user_id, None)


async def main() -> None:
    async with bot:
        await bot.start(TOKEN)


if __name__ == "__main__":
    asyncio.run(main())
